import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import io
from basic_inf import Basic_inf

class Show_graph():

    #保存先のファイルパス
    file = ''

    #グラフのフォント指定
    default_fontname = 'MS Gothic'
    default_fontsize = 12


    #店舗別総支払額の棒グラフを描画する関数
    def depict_purchase_amount(df: pd):

        fontname = Show_graph.default_fontname

        plt.rcParams["font.size"] = Show_graph.default_fontsize

        plt.xlabel('店舗',fontname=fontname)
        plt.ylabel('支払金額',fontname=fontname)
        plt.title('店舗別総支払金額',fontname=fontname)

        x = df['利用店舗']
        y = df['支払額']

        #文字が重なるので、店舗名を縦書きに変更
        plt.xticks(rotation=90,fontname=fontname)

        plt.bar(x, y,align='center')

        #グラフの間隔を調整
        plt.tight_layout()

        #Excelファイルのパスを取得
        file = Basic_inf.new_file_path

        #新しいタブを追加する
        workbook = openpyxl.load_workbook(file)
        new_sheet = workbook.create_sheet('店舗別総支払金額')

        #メモリに画像保存
        img_data = io.BytesIO()
        plt.savefig(img_data,format='png')

        img = openpyxl.drawing.image.Image(img_data)
        new_sheet.add_image(img,'A1')

        workbook.save(file)
        plt.clf()
        img_data.close()


    #店舗別利用回数の棒グラフを描画する関数
    def depict_purchase_frequency(df: pd):

        fontname = Show_graph.default_fontname

        plt.rcParams["font.size"] = Show_graph.default_fontsize

        plt.xlabel('店舗',fontname=fontname)
        plt.ylabel('利用回数',fontname=fontname)
        plt.title('店舗別利用回数',fontname=fontname)

        x = df['利用店舗']
        y = df['支払回数']

        #文字が重なるので、店舗名を縦書きに変更
        plt.xticks(rotation=90,fontname=fontname)

        plt.bar(x, y,align='center')

        #グラフの間隔を調整
        plt.tight_layout()

        file = Basic_inf.new_file_path

        workbook = openpyxl.load_workbook(file)
        new_sheet = workbook.create_sheet('店舗別利用回数')

        #メモリに画像保存
        img_data = io.BytesIO()
        plt.savefig(img_data,format='png')

        img = openpyxl.drawing.image.Image(img_data)
        new_sheet.add_image(img,'A1')

        workbook.save(file)
        plt.clf()
        img_data.close()

    #店舗別平均支払額の棒グラフを描画する関数
    def depict_mean_purchase(df: pd):

        fontname = Show_graph.default_fontname

        plt.rcParams["font.size"] = Show_graph.default_fontsize

        plt.xlabel('店舗',fontname=fontname)
        plt.ylabel('平均支払金額',fontname=fontname)
        plt.title('店舗別平均支払金額',fontname=fontname)

        x = df['利用店舗']
        y = df['平均支払金額']

        #文字が重なるので、店舗名を縦書きに変更
        plt.xticks(rotation=90,fontname=fontname)

        plt.bar(x, y,align='center')

        #グラフの間隔を調整
        plt.tight_layout()

        file = Basic_inf.new_file_path

        workbook = openpyxl.load_workbook(file)
        new_sheet = workbook.create_sheet('店舗別平均支払金額')

        #メモリに画像保存
        img_data = io.BytesIO()
        plt.savefig(img_data,format='png')

        img = openpyxl.drawing.image.Image(img_data)
        new_sheet.add_image(img,'A1')

        workbook.save(file)
        plt.clf()
        img_data.close()